import openpyxl

class readLoginExcel:
    def get_row_count(filename, sheetname):
        workbook = openpyxl.load_workbook(filename=filename)
        sheet = workbook[sheetname]
        return sheet.max_row

    def get_column_count(filename, sheetname):
        workbook = openpyxl.load_workbook(filename=filename)
        sheet = workbook[sheetname]
        return sheet.max_column

    def read_data(filename, sheetname, row, column):
        workbook = openpyxl.load_workbook(filename=filename)
        sheet = workbook[sheetname]
        return sheet.cell(row=row, column=column).value

    def write_data(filename, sheetname, row, column, data):
        workbook = openpyxl.load_workbook(filename=filename)
        sheet = workbook[sheetname]
        sheet.cell(row=row, column=column).value = data
        workbook.save(filename=filename)
    